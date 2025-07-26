import streamlit as st
import google.generativeai as genai
import json
import tempfile
import os
import re
import gspread
from google.oauth2.service_account import Credentials
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

st.set_page_config(page_title="PDF Extractor", layout="centered")
st.title("PDF Extractor")

with st.expander("Instructions"):
    st.markdown("""
    - Upload one or more PDF files using the uploader below.
    - Enter your Google Sheet ID or use the default one.
    - Click **Extract and Update** to process the uploaded PDFs and update Google Sheet.
    - The extracted data will also be available for download as a JSON file.
    """)

load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY")
genai.configure(api_key=API_KEY)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file"
]
CREDS_FILE = os.getenv("CREDS_JSON_ENV")
DEFAULT_SHEET_ID = '17tMHStXQYXaIQHQIA4jdUyHaYt_tuoNCEEuJCstWEuw'

COMPANY_NAME_ROW = 1
CONTACT_INFO_ROW = 2
HEADER_ROW = 3
ITEM_MASTER_LIST_COL = 2
COLUMNS_PER_SUPPLIER = 4

def extract_sheet_id_from_url(url):
    if not url:
        return None

    if "/" not in url and " " not in url and len(url) > 20:
        return url

    match = re.search(r'spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if match:
        return match.group(1)
            
    return None

prompt = """# System Message for Product List Extraction (PDF/Text Table Processing)

## Input Format
Provide PDF files (or images/tables with text extraction) containing product information (receipts, invoices, product lists, etc.):
* Upload PDF files or images directly
* Text will be extracted automatically
* Source should contain clear product details in readable text

## Task
Analyze the extracted text from documents and extract ALL product information to create a comprehensive product summary list.
Do not skip any products – ensure complete extraction of every line item, including physical consumables, product accessories, and items such as "ค่าเจาะรูกระจก" if they are charged per unit.

## CRITICAL: Complete Product Collection Rule
Extract ALL products as individual entries - quotations typically contain unique, detailed specifications.
- Each line item in the quotation represents a distinct product with specific details
- DO NOT consolidate or merge products - preserve all individual entries exactly as listed
- Products in quotations are already unique due to detailed specifications (sizes, locations, materials, etc.)
- Extract every single row/line item that has quantity, unit price, and total price
- Maintain the exact product descriptions including all specifications and location details

## Output Format (JSON only)
You must return ONLY this JSON structure:
{
  "company": "company name or first name + last name (NEVER null)",
  "vat": true,
  "name": "customer name or null",
  "contact": "phone number or email or null",
  "priceGuaranteeDay": 0,
  "products": [
    {
      "name": "full product description in Thai including all specifications AND size (EXCLUDE quantity/unit/price)",
      "quantity": 1,
      "unit": "match the unit shown in the pricePerUnit column (e.g., แผ่น, ตร.ม., ชิ้น, ตัว, เมตร, ชุด)",
      "pricePerUnit": 0,
      "totalPrice": 0
    }
  ],
  "totalPrice": 0,
  "totalVat": 0,
  "totalPriceIncludeVat": 0
}

## Field Extraction Guidelines

### name (Product Description)
* Include: material, model/type, size/dimensions, technical specs, finish, location details
* Include: ALL distinguishing characteristics that make each product unique
* Exclude: quantity, unit, price
* Preserve detailed specifications exactly as shown in quotation (sizes, locations, installation details)
* Example: "งานกระจกบานใส กระจกเทมเปอร์ เปรย์ เกร์ 10 มม. ฝังตัวยูเหล็ก สีเทา บันได ชั้น 1-2 ขนาด 0.975x4.672 ม."
* For items like "ค่าเจาะรูกระจก กว้าง 16มม.", include all distinguishing attributes in name

### unit and quantity (DIRECT EXTRACTION RULE)
* Extract unit and quantity DIRECTLY from each line item as shown
* Use the exact unit shown in the quotation (ชุด, แผ่น, ตร.ม., ชิ้น, ตัว, เมตร, etc.)
* Each line item represents a separate product entry - extract as individual entries
* If an item is priced per unit and has physical/product characteristics, treat as a product and include

### pricePerUnit and totalPrice
* If the document shows both product/material cost (ค่าวัสดุ) and labor/service cost (ค่าแรง) in the same row, always add the two (pricePerUnit = product per unit + service per unit)
* totalPrice must be calculated as: quantity × pricePerUnit
* Use numeric values only (no currency symbols)
* Extract cleanly from pricing fields as shown in each line item
* Allow minor rounding errors if visible in the document
* Extract each line item separately - do not combine or consolidate pricing
* If the quote has discount items, select only the discounted price

### CRITICAL: summaryItems and pricing summaries
* Extract ALL pricing summary items found at the end of the document into the summaryItems array
* Each summaryItem must have both a "label" and a "value" field
* Common labels to look for (extract EXACTLY as shown in document):
  - "รวม", "รวมเป็นเงิน", "ราคารวม", "Total", "TOTAL AMOUNT", "รวมราคา" - the initial subtotal
  - "ภาษีมูลค่าเพิ่ม 7%", "VAT 7%" - the VAT amount
  - "ค่าดำเนินการ 10%", "ค่าดำเนินการกำไร 12%", "Operating fee" - administrative fees
  - "ยอดรวมทั้งสิ้น", "รวมทั้งหมด", "รวมเงินทั้งสิน", "ราคารวมสุทธิ", "รวมราคางานทั้งหมดตามสัญญา", "TOTAL AMOUNT OF TENDER (INCLUDING VAT)" - the final total
* Extract the numeric values from each summary item exactly as shown (remove any commas, currency symbols)
* IMPORTANT: Also populate these specific fields:
  - totalPrice: Use the value from label matching "รวม", "รวมเป็นเงิน", "ราคารวม", "Total", etc. (subtotal before VAT)
  - totalVat: Use the value from label matching "ภาษีมูลค่าเพิ่ม 7%", "VAT 7%"
  - totalPriceIncludeVat: Use the value from label matching "ยอดรวมทั้งสิ้น", "รวมทั้งหมด", "รวมราคางานทั้งหมดตามสัญญา", etc. (final total)
* CRITICAL: All labels and numeric values must be extracted EXACTLY as shown in the document

## Complete Product Extraction Algorithm
1. Identify all line items with product descriptions, quantities, and prices
2. Extract each line item as a separate product - do not group or consolidate
3. Preserve all specifications and distinguishing details in the product name
4. Include location/installation details that make each product unique
5. Maintain individual quantities and pricing exactly as shown in quotation

## Inclusion/Exclusion Rules
* Extract ALL physical products and ALL items with price per unit, including consumables, parts, and accessories
* Include "ค่าเจาะรูกระจก" and similar items IF they are presented as a per-unit/consumable/physical item in the product table
* Exclude service/labor cost lines without a per-unit count (e.g., lump sum services)
* Each line item with specifications = separate product entry - preserve all individual entries
* Include all location-specific variations (different floors, units, areas) as separate products

## Quality Assurance Checklist
- [ ] ALL line items with products, quantities, and prices extracted
- [ ] Each product entry preserves detailed specifications and location details
- [ ] Product names include all distinguishing characteristics
- [ ] Quantities and prices match exactly what's shown in quotation
- [ ] No line items missed - complete extraction achieved
- [ ] All summaryItems accurately captured with exact labels and values
- [ ] totalPrice, totalVat, and totalPriceIncludeVat correctly mapped to appropriate summary values

## Final Notes
* PRIMARY GOAL: Extract every single product line item - quotations contain unique, detailed specifications
* Extract all line items meeting the above rules - preserve each as individual product entry
* Include all specs, sizes, and location details in name for complete product identification
* Match unit and quantity with pricePerUnit exactly as shown in quotation
* Maintain complete fidelity to the original quotation structure and details
* If the quote has discount items, select only the discounted price
* Ensure ALL summary pricing items are captured with exact labels and values
"""

matching_prompt = """
You're a product matching expert for construction materials in Thailand. Analyze products from List B against List A to find matches.

Matching criteria (in order of importance):
1. Material type (กระจก, อลูมิเนียม, เหล็ก, ไม้, etc.)
2. Product thickness (e.g., 10 มม., 12 มม.)
3. Product type/function (บานเลื่อน, บานสวิง, บานพับ, etc.)
4. Dimensions and measurements 
5. Location specifications

Consider these important rules:
- Two products match if they refer to the same physical item despite minor description variations
- Glass products must match thickness AND type exactly
- Different dimensions usually indicate different products
- Different locations (ชั้น 1, ชั้น 2, etc.) indicate different products
- Similar products with different finishes/colors are NOT matches

Return ONLY a JSON array of integers where each position corresponds to a product in List B:
- If a match exists in List A, return its index (0-based)
- If no match exists, return -1

List A:
{list_a}

List B:
{list_b}
"""

def extract_json_from_text(text):
    start_idx = text.find('{')
    end_idx = text.rfind('}') + 1
    if start_idx >= 0 and end_idx > start_idx:
        return json.loads(text[start_idx:end_idx])
    return None

def authenticate_and_open_sheet(sheet_id):
    creds = Credentials.from_service_account_file(CREDS_FILE, scopes=SCOPES)
    client = gspread.authorize(creds)
    spreadsheet = client.open_by_key(sheet_id)
    return spreadsheet.get_worksheet(0)

def ensure_first_three_rows_exist(worksheet):
    payloads = []
    for i in range(1, 4):
        payloads.append({
            'range': f"A{i}:B{i}",
            'values': [["", ""]]
        })
    if payloads:
        worksheet.batch_update(payloads, value_input_option='USER_ENTERED')

def find_next_available_column(worksheet):
    all_values = worksheet.get_all_values()
    max_col = ITEM_MASTER_LIST_COL  # Start with the item master list column
    
    if all_values:
        for row in all_values[:HEADER_ROW]:  # Check header rows only
            for i, cell in enumerate(row):
                if cell.strip():  # If cell has content
                    max_col = max(max_col, i + 1)  # +1 because sheets are 0-indexed but columns start at 1
    
    return max_col + 1  # Next available column

def match_products_with_gemini(existing_products, new_products):
    formatted_prompt = matching_prompt.format(
        list_a=json.dumps(existing_products, ensure_ascii=False),
        list_b=json.dumps(new_products, ensure_ascii=False)
    )
    
    response = genai.GenerativeModel(model_name="gemini-2.5-flash").generate_content(formatted_prompt)
    
    matches_text = response.text.strip()
    
    if '```' in matches_text:
        code_block_pattern = r'```(?:json)?(.*?)```'
        matches = re.findall(code_block_pattern, matches_text, re.DOTALL)
        if matches:
            matches_text = matches[0].strip()
    
    if matches_text.lower().startswith('json'):
        matches_text = matches_text[4:].strip()
    
    array_pattern = r'\[(.*?)\]'
    array_matches = re.search(array_pattern, matches_text, re.DOTALL)
    if array_matches:
        matches_text = f"[{array_matches.group(1)}]"
    
    if not matches_text:
        return [-1] * len(new_products)
    
    try:
        matches = json.loads(matches_text)
        if not isinstance(matches, list):
            matches = [-1] * len(new_products)
    except json.JSONDecodeError:
        matches = [-1] * len(new_products)
    
    if len(matches) < len(new_products):
        matches.extend([-1] * (len(new_products) - len(matches)))
    
    return matches

def check_sheet_template(worksheet):
    try:
        header_row = worksheet.row_values(HEADER_ROW)
        if header_row and len(header_row) >= 3 and "ปริมาณ" in header_row and "หน่วย" in header_row and "ราคาต่อหน่วย" in header_row:
            return True
        return False
    except:
        return False

def update_google_sheet_with_multiple_files(worksheet, all_json_data):
    ensure_first_three_rows_exist(worksheet)
    
    # Check if sheet template is valid
    has_valid_template = check_sheet_template(worksheet)
    if not has_valid_template:
        st.warning("Sheet template is not valid. Creating basic template.")
    
    # Get existing items and their metadata
    existing_items = worksheet.col_values(ITEM_MASTER_LIST_COL)[3:]
    existing_data = {}
    
    # Get all sheet data to analyze structure
    all_values = worksheet.get_all_values()
    if len(all_values) > HEADER_ROW:
        # Create a mapping of product names to their row indices
        for i, row in enumerate(all_values[HEADER_ROW:], start=HEADER_ROW+1):
            if len(row) > ITEM_MASTER_LIST_COL-1 and row[ITEM_MASTER_LIST_COL-1].strip():
                existing_data[row[ITEM_MASTER_LIST_COL-1]] = i
    
    start_row_index = 4
    payloads = []
    
    if all_json_data:
        # Find the next available column after the last used column
        next_col = find_next_available_column(worksheet)
        
        for idx, json_data in enumerate(all_json_data):
            supplier_col_start = next_col + (idx * COLUMNS_PER_SUPPLIER)
            
            # Add supplier info
            payloads.append({
                'range': f"{get_column_letter(supplier_col_start)}{COMPANY_NAME_ROW}",
                'values': [[json_data["company"]]]
            })
            
            payloads.append({
                'range': f"{get_column_letter(supplier_col_start)}{CONTACT_INFO_ROW}",
                'values': [[json_data["contact"] or ""]]
            })
            
            payloads.append({
                'range': f"{get_column_letter(supplier_col_start)}{HEADER_ROW}:{get_column_letter(supplier_col_start + 3)}{HEADER_ROW}",
                'values': [["ปริมาณ", "หน่วย", "ราคาต่อหน่วย", "รวมเป็นเงิน"]]
            })
            
            # Match products with existing ones
            with st.spinner(f"Matching products for supplier {idx+1}..."):
                new_product_names = [product["name"] for product in json_data["products"]]
                match_indices = match_products_with_gemini(existing_items, new_product_names)
                
                matched_products = []
                new_products = []
                
                for i, (product, match_idx) in enumerate(zip(json_data["products"], match_indices)):
                    if match_idx >= 0:
                        row_index = match_idx + start_row_index
                        matched_products.append((product, row_index))
                    else:
                        # Check if this product is in the master list by exact text match
                        product_name = product["name"]
                        if product_name in existing_data:
                            row_index = existing_data[product_name]
                            matched_products.append((product, row_index))
                        else:
                            new_products.append(product)
            
            # Update matched products
            for product, row_index in matched_products:
                payloads.append({
                    'range': f"{get_column_letter(supplier_col_start)}{row_index}:{get_column_letter(supplier_col_start + 3)}{row_index}",
                    'values': [[product["quantity"], product["unit"], product["pricePerUnit"], product["totalPrice"]]]
                })
            
            # Add new products
            for product in new_products:
                product_name = product["name"]
                existing_items.append(product_name)
                existing_data[product_name] = len(existing_items) + start_row_index - 1
                row_index = existing_data[product_name]
                
                payloads.append({
                    'range': f"B{row_index}",
                    'values': [[product_name]]
                })
                
                payloads.append({
                    'range': f"{get_column_letter(supplier_col_start)}{row_index}:{get_column_letter(supplier_col_start + 3)}{row_index}",
                    'values': [[product["quantity"], product["unit"], product["pricePerUnit"], product["totalPrice"]]]
                })
            
            # Add summary rows
            summary_start_row = max(existing_data.values()) + 2 if existing_data else start_row_index + len(existing_items) + 1
            
            payloads.append({
                'range': f"B{summary_start_row}",
                'values': [["รวมเป็นเงิน"]]
            })
            payloads.append({
                'range': f"{get_column_letter(supplier_col_start + 3)}{summary_start_row}",
                'values': [[json_data["totalPrice"]]]
            })
            
            summary_start_row += 1
            payloads.append({
                'range': f"B{summary_start_row}",
                'values': [["ภาษีมูลค่าเพิ่ม 7%"]]
            })
            payloads.append({
                'range': f"{get_column_letter(supplier_col_start + 3)}{summary_start_row}",
                'values': [[json_data["totalVat"]]]
            })
            
            summary_start_row += 1
            payloads.append({
                'range': f"B{summary_start_row}",
                'values': [["ยอดรวมทั้งสิ้น"]]
            })
            payloads.append({
                'range': f"{get_column_letter(supplier_col_start + 3)}{summary_start_row}",
                'values': [[json_data["totalPriceIncludeVat"]]]
            })
            
            if "match_stats" not in st.session_state:
                st.session_state.match_stats = {}
            
            st.session_state.match_stats[idx] = {
                "matched": len(matched_products),
                "new": len(new_products),
                "total": len(json_data["products"])
            }

    if payloads:
        worksheet.batch_update(payloads, value_input_option='USER_ENTERED')
    
    return len(all_json_data)

uploaded_files = st.file_uploader(
    "Upload one or more PDF files", 
    type=["pdf"], 
    accept_multiple_files=True
)

sheet_id = st.text_input("Google Sheet ID", value=DEFAULT_SHEET_ID)

if st.button("Extract and Update") and uploaded_files and sheet_id:
    all_data = []
    progress = st.progress(0)
    
    worksheet = authenticate_and_open_sheet(sheet_id)
    st.info("Connected to Google Sheet successfully.")
    
    for idx, uploaded_file in enumerate(uploaded_files):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
            tmp_file.write(uploaded_file.read())
            tmp_file_path = tmp_file.name
            
        with st.spinner(f"Processing {uploaded_file.name}..."):
            gemini_file = genai.upload_file(path=tmp_file_path, display_name=f"PDF for Extraction: {uploaded_file.name}")
            response = genai.GenerativeModel(model_name="gemini-2.5-flash").generate_content([prompt, gemini_file])
            
            json_data = extract_json_from_text(response.text)
            if json_data:
                all_data.append(json_data)
                st.success(f"Extracted data from {uploaded_file.name}")
            else:
                st.warning(f"Failed to extract data from {uploaded_file.name}")
                
        os.unlink(tmp_file_path)
        progress.progress((idx + 1) / len(uploaded_files))

    if all_data:
        with st.spinner("Updating Google Sheet with AI product matching..."):
            suppliers_updated = update_google_sheet_with_multiple_files(worksheet, all_data)
            st.success(f"Updated Google Sheet with data from {suppliers_updated} supplier(s)")
            
            if "match_stats" in st.session_state:
                for idx, stats in st.session_state.match_stats.items():
                    st.info(f"Supplier {idx+1} product matching: {stats['matched']} products matched, {stats['new']} new products added (total: {stats['total']})")
        
        merged_data = {"extractions": all_data}
        json_str = json.dumps(merged_data, indent=2, ensure_ascii=False)
        
        st.download_button(
            label="Download Extracted JSON File",
            data=json_str,
            file_name="extracted_data.json",
            mime="application/json"
        )
        
        st.subheader("Extracted Data")
        st.json(json_str)